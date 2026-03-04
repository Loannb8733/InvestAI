"""Tests for report generation service."""

import io
from datetime import datetime
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from app.services.report_service import AssetReport, PortfolioSummary, ReportService, TaxSummary2086, TaxTransaction


@pytest.fixture
def report_service():
    return ReportService()


@pytest.fixture
def sample_data():
    """Sample portfolio data for report generation."""
    return {
        "portfolios": [
            PortfolioSummary(
                name="Crypto",
                total_value=15000.0,
                total_invested=10000.0,
                gain_loss=5000.0,
                gain_loss_percent=50.0,
                asset_count=3,
            ),
            PortfolioSummary(
                name="Actions",
                total_value=8000.0,
                total_invested=7500.0,
                gain_loss=500.0,
                gain_loss_percent=6.67,
                asset_count=2,
            ),
        ],
        "assets": [
            AssetReport(
                symbol="BTC",
                name="Bitcoin",
                asset_type="crypto",
                quantity=0.5,
                avg_buy_price=20000.0,
                current_price=30000.0,
                total_invested=10000.0,
                current_value=15000.0,
                gain_loss=5000.0,
                gain_loss_percent=50.0,
            ),
            AssetReport(
                symbol="AAPL",
                name="Apple",
                asset_type="stock",
                quantity=10.0,
                avg_buy_price=150.0,
                current_price=180.0,
                total_invested=1500.0,
                current_value=1800.0,
                gain_loss=300.0,
                gain_loss_percent=20.0,
            ),
        ],
        "transactions": [
            TaxTransaction(
                date=datetime(2025, 3, 15),
                symbol="BTC",
                transaction_type="sell",
                quantity=0.1,
                price=28000.0,
                total=2800.0,
                fee=5.0,
                gain_loss=800.0,
            ),
        ],
        "summary": {
            "total_value": 23000.0,
            "total_invested": 17500.0,
            "gain_loss": 5500.0,
            "gain_loss_percent": 31.43,
        },
        "generated_at": datetime(2025, 6, 1),
        "year": 2025,
    }


@pytest.fixture
def empty_data():
    return {
        "portfolios": [],
        "assets": [],
        "transactions": [],
        "summary": {
            "total_value": 0,
            "total_invested": 0,
            "gain_loss": 0,
            "gain_loss_percent": 0,
        },
    }


class TestPerformancePDF:
    """Tests for PDF performance report generation."""

    def test_generates_bytes(self, report_service, sample_data):
        result = report_service.generate_performance_pdf(sample_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_pdf_header(self, report_service, sample_data):
        result = report_service.generate_performance_pdf(sample_data)
        assert result[:5] == b"%PDF-"

    def test_empty_data_still_generates(self, report_service, empty_data):
        result = report_service.generate_performance_pdf(empty_data)
        assert isinstance(result, bytes)
        assert result[:5] == b"%PDF-"


class TestPerformanceExcel:
    """Tests for Excel performance report generation."""

    def test_generates_bytes(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        assert "Résumé" in wb.sheetnames

    def test_has_portfolio_sheet(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        assert "Portefeuilles" in wb.sheetnames

    def test_has_assets_sheet(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        assert "Actifs" in wb.sheetnames

    def test_has_transactions_sheet(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        assert "Transactions" in wb.sheetnames

    def test_summary_values(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Résumé"]
        assert ws["A5"].value == "Valeur totale"
        assert ws["B5"].value == 23000.0

    def test_empty_data_no_extra_sheets(self, report_service, empty_data):
        result = report_service.generate_performance_excel(empty_data)
        wb = load_workbook(io.BytesIO(result))
        assert "Résumé" in wb.sheetnames
        # No portfolios/assets/transactions sheets when empty
        assert "Portefeuilles" not in wb.sheetnames

    def test_asset_data_in_sheet(self, report_service, sample_data):
        result = report_service.generate_performance_excel(sample_data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Actifs"]
        # Row 2 should have BTC data
        assert ws.cell(row=2, column=1).value == "BTC"


class TestTaxExcel:
    """Tests for tax report Excel generation."""

    @pytest.fixture
    def tax_summary_with_cession(self):
        """TaxSummary2086 with 1 cession."""
        return TaxSummary2086(
            year=2025,
            total_cessions=15000.0,
            total_acquisitions_fraction=10000.0,
            total_plus_values=5000.0,
            total_moins_values=0.0,
            net_plus_value=5000.0,
            nb_cessions=1,
            nb_court_terme=1,
            nb_long_terme=0,
            flat_tax_30=1500.0,
            ir_12_8=640.0,
            ps_17_2=860.0,
            events=[],
        )

    @pytest.fixture
    def tax_summary_empty(self):
        """TaxSummary2086 with 0 cessions."""
        return TaxSummary2086(
            year=2025,
            total_cessions=0.0,
            total_acquisitions_fraction=0.0,
            total_plus_values=0.0,
            total_moins_values=0.0,
            net_plus_value=0.0,
            nb_cessions=0,
            nb_court_terme=0,
            nb_long_terme=0,
            flat_tax_30=0.0,
            ir_12_8=0.0,
            ps_17_2=0.0,
            events=[],
        )

    @pytest.mark.asyncio
    async def test_generates_bytes(self, report_service, tax_summary_with_cession):
        with patch.object(
            report_service, "compute_tax_2086", new_callable=AsyncMock, return_value=tax_summary_with_cession
        ):
            result = await report_service.generate_tax_excel(None, "test_user", 2025)
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_sheet_title_contains_year(self, report_service, tax_summary_with_cession):
        with patch.object(
            report_service, "compute_tax_2086", new_callable=AsyncMock, return_value=tax_summary_with_cession
        ):
            result = await report_service.generate_tax_excel(None, "test_user", 2025)
        wb = load_workbook(io.BytesIO(result))
        assert "Résumé 2025" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_transaction_count(self, report_service, tax_summary_with_cession):
        with patch.object(
            report_service, "compute_tax_2086", new_callable=AsyncMock, return_value=tax_summary_with_cession
        ):
            result = await report_service.generate_tax_excel(None, "test_user", 2025)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # B5 = nombre de cessions
        assert ws["B5"].value == 1

    @pytest.mark.asyncio
    async def test_empty_transactions(self, report_service, tax_summary_empty):
        with patch.object(report_service, "compute_tax_2086", new_callable=AsyncMock, return_value=tax_summary_empty):
            result = await report_service.generate_tax_excel(None, "test_user", 2025)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws["B5"].value == 0  # 0 cessions
