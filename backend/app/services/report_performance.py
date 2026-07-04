"""Performance report rendering (PDF + Excel), extracted from report_service."""

from __future__ import annotations

import io
import logging
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

from app.services.report_common import (
    _BLUE,
    _BORDER_COLOR,
    _DARK_BLUE,
    _LIGHT_BG,
    _XL_BORDER,
    _XL_HEADER_FILL,
    _XL_HEADER_FONT,
    _XL_MONEY,
    _XL_PERCENT,
    _build_doc_with_footer,
    _fmt_qty,
    _format_paris_dt,
    _gain_color,
    _money,
    _pct,
)

logger = logging.getLogger(__name__)


class PerformanceReportMixin:
    """Mixed into ReportService."""

    def generate_performance_pdf(self, data: Dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = _build_doc_with_footer(
            buffer,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("T", parent=styles["Heading1"], fontSize=24, spaceAfter=30, textColor=_BLUE)
        heading_style = ParagraphStyle(
            "H",
            parent=styles["Heading2"],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=_DARK_BLUE,
        )
        normal_style = styles["Normal"]
        elements = []

        elements.append(Paragraph("Rapport de Performance", title_style))
        elements.append(Paragraph(f"Généré le {_format_paris_dt()}", normal_style))
        elements.append(Spacer(1, 20))

        summary = data.get("summary", {})
        pnl = data.get("pnl_data", {})
        elements.append(Paragraph("Résumé Global", heading_style))
        summary_data = [
            ["Indicateur", "Valeur"],
            ["Valeur totale", _money(summary.get("total_value", 0))],
            ["Total investi", _money(summary.get("total_invested", 0))],
            ["Plus/Moins-value", _money(summary.get("gain_loss", 0))],
            ["Performance", _pct(summary.get("gain_loss_percent", 0))],
        ]
        if pnl:
            summary_data.extend(
                [
                    ["", ""],
                    ["P&L Latent", _money(pnl.get("unrealized_pnl", 0))],
                    ["P&L Réalisé", _money(pnl.get("realized_pnl", 0))],
                    ["Total Frais", _money(pnl.get("total_fees", 0))],
                    ["P&L Net", _money(pnl.get("net_pnl", 0))],
                ]
            )
        t = Table(summary_data, colWidths=[8 * cm, 6 * cm])
        summary_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), _LIGHT_BG),
            ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
        ]
        if pnl:
            # Bold + color on P&L Net row (last row)
            net_pnl = pnl.get("net_pnl", 0)
            summary_style_cmds.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
            summary_style_cmds.append(("TEXTCOLOR", (1, -1), (1, -1), _gain_color(net_pnl)))
        t.setStyle(TableStyle(summary_style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Attribution de performance
        attr = data.get("attribution", {})
        if any(v > 0 for v in attr.values()):
            elements.append(Paragraph("Attribution de Performance", heading_style))
            total_attr = sum(attr.values()) or 1
            attr_data = [
                ["Catégorie", "Valeur", "Poids"],
                [
                    "Alpha (Altcoins)",
                    _money(attr.get("alpha", 0)),
                    _pct(attr.get("alpha", 0) / total_attr * 100),
                ],
                [
                    "Beta (BTC)",
                    _money(attr.get("beta", 0)),
                    _pct(attr.get("beta", 0) / total_attr * 100),
                ],
                [
                    "Protection (Or)",
                    _money(attr.get("protection", 0)),
                    _pct(attr.get("protection", 0) / total_attr * 100),
                ],
                [
                    "Munitions (Cash/Stables)",
                    _money(attr.get("munitions", 0)),
                    _pct(attr.get("munitions", 0) / total_attr * 100),
                ],
            ]
            t = Table(attr_data, colWidths=[6 * cm, 4 * cm, 4 * cm])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), _LIGHT_BG),
                    ]
                )
            )
            elements.append(t)
            elements.append(Spacer(1, 20))

        # Flux de trésorerie (Cash Flows)
        cf = data.get("cash_flows", {})
        if cf.get("total_deposits", 0) > 0 or cf.get("total_withdrawals", 0) > 0:
            elements.append(Paragraph("Flux de Trésorerie", heading_style))
            gain = summary.get("gain_loss", 0)
            cf_data = [
                ["Flux", "Montant"],
                ["Total déposé", _money(cf.get("total_deposits", 0))],
                ["Total retiré", _money(cf.get("total_withdrawals", 0))],
                ["Flux nets (dépôts − retraits)", _money(cf.get("net_flows", 0))],
                ["Plus/Moins-value réelle", _money(gain)],
            ]
            t = Table(cf_data, colWidths=[8 * cm, 6 * cm])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), _LIGHT_BG),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("TEXTCOLOR", (1, -1), (1, -1), _gain_color(gain)),
                    ]
                )
            )
            elements.append(t)
            elements.append(Spacer(1, 20))

        # Portfolios
        portfolios = data.get("portfolios", [])
        if portfolios:
            elements.append(Paragraph("Portefeuilles", heading_style))
            rows = [["Nom", "Valeur", "Investi", "+/- Value", "Perf.", "Actifs"]]
            for p in portfolios:
                rows.append(
                    [
                        p.name,
                        _money(p.total_value),
                        _money(p.total_invested),
                        _money(p.gain_loss),
                        _pct(p.gain_loss_percent),
                        str(p.asset_count),
                    ]
                )
            t = Table(rows, colWidths=[4 * cm, 2.8 * cm, 2.8 * cm, 2.5 * cm, 2 * cm, 1.5 * cm])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_BG]),
                    ]
                )
            )
            elements.append(t)
            elements.append(Spacer(1, 20))

        # Assets
        assets = data.get("assets", [])
        if assets:
            elements.append(Paragraph("Détail des Actifs", heading_style))
            rows = [
                [
                    "Symbole",
                    "Type",
                    "Qté",
                    "PRU",
                    "Valeur",
                    "+/- Value",
                    "Perf.",
                    "Break-even",
                    "Risque",
                ]
            ]
            for a in assets:
                rows.append(
                    [
                        a.symbol,
                        a.asset_type.upper()[:6],
                        _fmt_qty(a.quantity),
                        _money(a.avg_buy_price),
                        _money(a.current_value),
                        _money(a.gain_loss),
                        _pct(a.gain_loss_percent),
                        _money(a.breakeven_price) if a.breakeven_price else "N/A",
                        f"{a.risk_weight:.1f}%" if a.risk_weight else "N/A",
                    ]
                )
            t = Table(
                rows,
                colWidths=[
                    2.0 * cm,
                    1.5 * cm,
                    1.6 * cm,
                    2.0 * cm,
                    2.2 * cm,
                    2.0 * cm,
                    1.6 * cm,
                    2.0 * cm,
                    1.5 * cm,
                ],
            )
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_BG]),
                    ]
                )
            )
            elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Performance Excel ───────────────────────────────────────────

    def generate_performance_excel(self, data: Dict[str, Any]) -> bytes:
        wb = Workbook()

        # ── Résumé sheet ─────────────────────────────────────────────
        ws = wb.active
        ws.title = "Résumé"
        summary = data.get("summary", {})
        pnl = data.get("pnl_data", {})
        ws["A1"] = "Rapport de Performance InvestAI"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Généré le {_format_paris_dt()}"

        for cell_ref in ["A4", "B4"]:
            ws[cell_ref].font = _XL_HEADER_FONT
            ws[cell_ref].fill = _XL_HEADER_FILL
        ws["A4"] = "Indicateur"
        ws["B4"] = "Valeur"

        summary_rows = [
            ("Valeur totale", summary.get("total_value", 0), _XL_MONEY),
            ("Total investi", summary.get("total_invested", 0), _XL_MONEY),
            ("Plus/Moins-value", summary.get("gain_loss", 0), _XL_MONEY),
            ("Performance (%)", summary.get("gain_loss_percent", 0) / 100, _XL_PERCENT),
        ]
        if pnl:
            summary_rows.extend(
                [
                    ("", "", None),
                    ("P&L Latent", pnl.get("unrealized_pnl", 0), _XL_MONEY),
                    ("P&L Réalisé", pnl.get("realized_pnl", 0), _XL_MONEY),
                    ("Total Frais", pnl.get("total_fees", 0), _XL_MONEY),
                    ("P&L Net", pnl.get("net_pnl", 0), _XL_MONEY),
                ]
            )
        for i, (label, value, fmt) in enumerate(summary_rows, start=5):
            ws[f"A{i}"] = label
            if value != "":
                ws[f"B{i}"] = value
            if fmt:
                ws[f"B{i}"].number_format = fmt
        # Bold P&L Net row
        if pnl:
            net_row = 5 + len(summary_rows) - 1
            ws[f"A{net_row}"].font = Font(bold=True)
            ws[f"B{net_row}"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18

        # ── Portefeuilles sheet ──────────────────────────────────────
        portfolios = data.get("portfolios", [])
        if portfolios:
            ws2 = wb.create_sheet("Portefeuilles")
            headers = [
                "Nom",
                "Valeur",
                "Investi",
                "+/- Value",
                "Performance",
                "Nb Actifs",
            ]
            for col, h in enumerate(headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = _XL_HEADER_FONT
                c.fill = _XL_HEADER_FILL
                c.border = _XL_BORDER
            for row, p in enumerate(portfolios, 2):
                ws2.cell(row=row, column=1, value=p.name).border = _XL_BORDER
                ws2.cell(row=row, column=2, value=p.total_value).number_format = _XL_MONEY
                ws2.cell(row=row, column=3, value=p.total_invested).number_format = _XL_MONEY
                ws2.cell(row=row, column=4, value=p.gain_loss).number_format = _XL_MONEY
                ws2.cell(row=row, column=5, value=p.gain_loss_percent / 100).number_format = _XL_PERCENT
                ws2.cell(row=row, column=6, value=p.asset_count)
                for col in range(1, 7):
                    ws2.cell(row=row, column=col).border = _XL_BORDER
            for col in range(1, 7):
                ws2.column_dimensions[get_column_letter(col)].width = 15

        # ── Actifs sheet (enrichi) ───────────────────────────────────
        assets = data.get("assets", [])
        if assets:
            ws3 = wb.create_sheet("Actifs")
            headers = [
                "Symbole",
                "Nom",
                "Type",
                "Plateforme",
                "Quantité",
                "PRU",
                "Prix actuel",
                "Break-even",
                "Investi",
                "Valeur",
                "+/- Value",
                "Perf.",
                "Frais",
                "Risque",
            ]
            for col, h in enumerate(headers, 1):
                c = ws3.cell(row=1, column=col, value=h)
                c.font = _XL_HEADER_FONT
                c.fill = _XL_HEADER_FILL
                c.border = _XL_BORDER
            money_cols = {
                6,
                7,
                8,
                9,
                10,
                11,
                13,
            }  # PRU, Prix, Break-even, Investi, Valeur, +/-, Frais
            pct_cols = {12, 14}  # Perf., Risque
            for row, a in enumerate(assets, 2):
                vals = [
                    a.symbol,
                    a.name,
                    a.asset_type,
                    a.exchange or "",
                    a.quantity,
                    a.avg_buy_price,
                    a.current_price,
                    a.breakeven_price if a.breakeven_price else "",
                    a.total_invested,
                    a.current_value,
                    a.gain_loss,
                    a.gain_loss_percent / 100,
                    a.total_fees,
                    a.risk_weight / 100 if a.risk_weight else "",
                ]
                for col, v in enumerate(vals, 1):
                    c = ws3.cell(row=row, column=col, value=v)
                    c.border = _XL_BORDER
                    if col in money_cols and v != "":
                        c.number_format = _XL_MONEY
                    elif col in pct_cols and v != "":
                        c.number_format = _XL_PERCENT
            for col in range(1, len(headers) + 1):
                ws3.column_dimensions[get_column_letter(col)].width = 14

        # ── Transactions sheet ───────────────────────────────────────
        transactions = data.get("transactions", [])
        if transactions:
            ws4 = wb.create_sheet("Transactions")
            headers = ["Date", "Symbole", "Type", "Quantité", "Prix", "Total", "Frais"]
            for col, h in enumerate(headers, 1):
                c = ws4.cell(row=1, column=col, value=h)
                c.font = _XL_HEADER_FONT
                c.fill = _XL_HEADER_FILL
                c.border = _XL_BORDER
            for row, t in enumerate(transactions, 2):
                date_str = t.date.strftime("%d/%m/%Y") if t.date else ""
                ws4.cell(row=row, column=1, value=date_str).border = _XL_BORDER
                ws4.cell(row=row, column=2, value=t.symbol).border = _XL_BORDER
                ws4.cell(row=row, column=3, value=t.transaction_type).border = _XL_BORDER
                ws4.cell(row=row, column=4, value=t.quantity).border = _XL_BORDER
                ws4.cell(row=row, column=5, value=t.price).number_format = _XL_MONEY
                ws4.cell(row=row, column=6, value=t.total).number_format = _XL_MONEY
                ws4.cell(row=row, column=7, value=t.fee).number_format = _XL_MONEY
                for col in range(1, 8):
                    ws4.cell(row=row, column=col).border = _XL_BORDER
            for col in range(1, 8):
                ws4.column_dimensions[get_column_letter(col)].width = 14

        # ── Analyse de Risque sheet ──────────────────────────────────
        risk = data.get("risk_metrics", {})
        if risk:
            ws_risk = wb.create_sheet("Analyse de Risque")
            for cell_ref in ["A1", "B1"]:
                ws_risk[cell_ref].font = _XL_HEADER_FONT
                ws_risk[cell_ref].fill = _XL_HEADER_FILL
                ws_risk[cell_ref].border = _XL_BORDER
            ws_risk["A1"] = "Indicateur"
            ws_risk["B1"] = "Valeur"

            mdd = risk.get("max_drawdown", {})
            var95 = risk.get("var_95", {})
            conc = risk.get("concentration", {})
            st20 = risk.get("stress_test_20", {})
            st40 = risk.get("stress_test_40", {})

            risk_rows = [
                ("Volatilité annualisée", risk.get("volatility", 0) / 100, _XL_PERCENT),
                ("Ratio de Sharpe", risk.get("sharpe_ratio", 0), "0.00"),
                ("Max Drawdown", mdd.get("max_drawdown_percent", 0) / 100, _XL_PERCENT),
                ("Max Drawdown — Pic", mdd.get("peak_date", ""), None),
                ("Max Drawdown — Creux", mdd.get("trough_date", ""), None),
                ("VaR 95% (%)", var95.get("var_percent", 0) / 100, _XL_PERCENT),
                ("VaR 95% (montant)", var95.get("var_amount", 0), _XL_MONEY),
                ("", "", None),
                ("Indice HHI", conc.get("hhi", 0), "0"),
                ("Concentration", conc.get("interpretation", ""), None),
                ("Actif dominant", conc.get("top_asset", ""), None),
                (
                    "Poids actif dominant",
                    (conc.get("top_concentration", 0) or 0) / 100,
                    _XL_PERCENT,
                ),
                ("", "", None),
                ("Stress Test -20% (perte)", st20.get("potential_loss", 0), _XL_MONEY),
                ("Stress Test -20% (valeur)", st20.get("stressed_value", 0), _XL_MONEY),
                ("Stress Test -40% (perte)", st40.get("potential_loss", 0), _XL_MONEY),
                ("Stress Test -40% (valeur)", st40.get("stressed_value", 0), _XL_MONEY),
            ]
            for i, (label, value, fmt) in enumerate(risk_rows, start=2):
                ws_risk[f"A{i}"] = label
                ws_risk[f"A{i}"].border = _XL_BORDER
                if value != "":
                    ws_risk[f"B{i}"] = value
                ws_risk[f"B{i}"].border = _XL_BORDER
                if fmt:
                    ws_risk[f"B{i}"].number_format = fmt

            ws_risk.column_dimensions["A"].width = 28
            ws_risk.column_dimensions["B"].width = 18

        # ── Attribution sheet ──────────────────────────────────────────
        attr = data.get("attribution", {})
        if any(v > 0 for v in attr.values()):
            ws_attr = wb.create_sheet("Attribution")
            ws_attr["A1"] = "Attribution de Performance"
            ws_attr["A1"].font = Font(bold=True, size=14)

            attr_headers = ["Catégorie", "Valeur", "Poids"]
            for col, h in enumerate(attr_headers, 1):
                c = ws_attr.cell(row=3, column=col, value=h)
                c.font = _XL_HEADER_FONT
                c.fill = _XL_HEADER_FILL
                c.border = _XL_BORDER

            total_attr = sum(attr.values()) or 1
            attr_rows = [
                ("Alpha (Altcoins)", attr.get("alpha", 0)),
                ("Beta (BTC)", attr.get("beta", 0)),
                ("Protection (Or)", attr.get("protection", 0)),
                ("Munitions (Cash/Stables)", attr.get("munitions", 0)),
            ]
            for i, (label, value) in enumerate(attr_rows, 4):
                ws_attr.cell(row=i, column=1, value=label).border = _XL_BORDER
                c_val = ws_attr.cell(row=i, column=2, value=value)
                c_val.number_format = _XL_MONEY
                c_val.border = _XL_BORDER
                c_pct = ws_attr.cell(row=i, column=3, value=value / total_attr)
                c_pct.number_format = _XL_PERCENT
                c_pct.border = _XL_BORDER

            ws_attr.column_dimensions["A"].width = 28
            ws_attr.column_dimensions["B"].width = 16
            ws_attr.column_dimensions["C"].width = 12

        # ── Flux de Trésorerie sheet ──────────────────────────────────
        cf = data.get("cash_flows", {})
        if cf.get("total_deposits", 0) > 0 or cf.get("total_withdrawals", 0) > 0:
            ws_cf = wb.create_sheet("Flux de Trésorerie")
            ws_cf["A1"] = "Flux de Trésorerie (Cash Flows)"
            ws_cf["A1"].font = Font(bold=True, size=14)

            for cell_ref in ["A3", "B3"]:
                ws_cf[cell_ref].font = _XL_HEADER_FONT
                ws_cf[cell_ref].fill = _XL_HEADER_FILL
                ws_cf[cell_ref].border = _XL_BORDER
            ws_cf["A3"] = "Flux"
            ws_cf["B3"] = "Montant"

            cf_rows = [
                ("Total déposé", cf.get("total_deposits", 0)),
                ("Total retiré", cf.get("total_withdrawals", 0)),
                ("Flux nets (dépôts − retraits)", cf.get("net_flows", 0)),
                ("Plus/Moins-value réelle", summary.get("gain_loss", 0)),
            ]
            for i, (label, value) in enumerate(cf_rows, 4):
                ws_cf.cell(row=i, column=1, value=label).border = _XL_BORDER
                c = ws_cf.cell(row=i, column=2, value=value)
                c.number_format = _XL_MONEY
                c.border = _XL_BORDER

            ws_cf.column_dimensions["A"].width = 30
            ws_cf.column_dimensions["B"].width = 18

        # ── Performance par Plateforme sheet ─────────────────────────
        platforms = data.get("platform_analysis", [])
        if platforms:
            ws_plat = wb.create_sheet("Perf. par Plateforme")
            plat_headers = [
                "Plateforme",
                "Nb Actifs",
                "Valeur",
                "Investi",
                "Frais",
                "P&L Net",
                "ROI %",
            ]
            for col, h in enumerate(plat_headers, 1):
                c = ws_plat.cell(row=1, column=col, value=h)
                c.font = _XL_HEADER_FONT
                c.fill = _XL_HEADER_FILL
                c.border = _XL_BORDER
            for row, p in enumerate(platforms, 2):
                vals = [
                    p["exchange"],
                    p["count"],
                    p["total_value"],
                    p["total_invested"],
                    p["total_fees"],
                    p["net_pnl"],
                    p["roi_percent"] / 100,
                ]
                for col, v in enumerate(vals, 1):
                    c = ws_plat.cell(row=row, column=col, value=v)
                    c.border = _XL_BORDER
                    if col in (3, 4, 5, 6):
                        c.number_format = _XL_MONEY
                    elif col == 7:
                        c.number_format = _XL_PERCENT
            for col in range(1, len(plat_headers) + 1):
                ws_plat.column_dimensions[get_column_letter(col)].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Tax 2086 PDF ────────────────────────────────────────────────
