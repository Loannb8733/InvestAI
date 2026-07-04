"""Report generation service for PDF and Excel exports."""

import io
import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.portfolio import Portfolio
from app.models.transaction import Transaction
from app.services.asset_classification import is_liquidity, is_safe_haven
from app.services.metrics_service import metrics_service
from app.services.report_common import (
    _BLUE,
    _BORDER_COLOR,
    _DARK_BLUE,
    _GREEN,
    _LIGHT_BG,
    _RED,
    _XL_BORDER,
    _XL_HEADER_FILL,
    _XL_HEADER_FONT,
    _XL_MONEY,
    _build_doc_with_footer,
    _money,
    _now_paris,
)
from app.services.report_performance import PerformanceReportMixin
from app.services.report_rebalancing import (  # noqa: F401 — re-exported
    RebalanceOrder,
    RebalancingReport,
    RebalancingReportMixin,
)
from app.services.report_tax import (  # noqa: F401 — re-exported for back-compat
    TAX_MODE,
    TAX_MODE_LEGACY,
    TaxComputeMixin,
    TaxEvent2086,
    TaxSummary2086,
    _TaxMode,
)
from app.services.report_transactions import TransactionsReportMixin
from app.services.snapshot_service import snapshot_service

logger = logging.getLogger(__name__)

# ── Data classes ────────────────────────────────────────────────────


@dataclass
class PortfolioSummary:
    name: str
    total_value: float
    total_invested: float
    gain_loss: float
    gain_loss_percent: float
    asset_count: int


@dataclass
class AssetReport:
    symbol: str
    name: str
    asset_type: str
    quantity: float
    avg_buy_price: float
    current_price: float
    total_invested: float
    current_value: float
    gain_loss: float
    gain_loss_percent: float
    breakeven_price: Optional[float] = None
    risk_weight: Optional[float] = None
    total_fees: float = 0.0
    exchange: Optional[str] = None


@dataclass
class TaxTransaction:
    date: datetime
    symbol: str
    transaction_type: str
    quantity: float
    price: float
    total: float
    fee: float
    gain_loss: Optional[float]


# ── Crypto asset class mapping (server-side, mirrors frontend) ─────


# ── Report Service ──────────────────────────────────────────────────


class ReportService(
    TaxComputeMixin,
    RebalancingReportMixin,
    PerformanceReportMixin,
    TransactionsReportMixin,
):
    """Service for generating various reports."""

    # ── Report data (single source of truth: metrics_service) ──────

    async def get_report_data(
        self,
        db: AsyncSession,
        user_id: str,
        year: Optional[int] = None,
        currency: str = "EUR",
    ) -> Dict[str, Any]:
        """Get comprehensive portfolio data for reports.

        Delegates ALL calculations to metrics_service and snapshot_service.
        This method only reshapes their output into report-friendly structures.
        """
        # 1. Dashboard-level summary (includes pnl_data, allocation)
        dashboard = await metrics_service.get_user_dashboard_metrics(db, user_id, currency=currency)

        # 2. Per-portfolio metrics for detailed asset breakdown
        result = await db.execute(select(Portfolio).where(Portfolio.user_id == user_id))
        portfolios = result.scalars().all()

        if not portfolios:
            return {
                "portfolios": [],
                "assets": [],
                "transactions": [],
                "summary": {
                    "total_value": 0.0,
                    "total_invested": 0.0,
                    "gain_loss": 0.0,
                    "gain_loss_percent": 0.0,
                },
                "pnl_data": {},
                "risk_metrics": {},
                "platform_analysis": [],
                "generated_at": datetime.now(timezone.utc),
                "year": year,
            }

        portfolio_summaries: List[PortfolioSummary] = []
        all_asset_reports: List[AssetReport] = []
        all_asset_ids: List[str] = []

        for portfolio in portfolios:
            pm = await metrics_service.get_portfolio_metrics(db, str(portfolio.id), currency=currency)

            p_value = pm["total_value"]
            p_invested = pm["total_invested"]
            p_gain = p_value - p_invested
            p_gain_pct = (p_gain / p_invested * 100) if p_invested > 0 else 0

            portfolio_summaries.append(
                PortfolioSummary(
                    name=portfolio.name,
                    total_value=p_value,
                    total_invested=p_invested,
                    gain_loss=p_gain,
                    gain_loss_percent=p_gain_pct,
                    asset_count=pm.get("assets_count", len(pm.get("assets", []))),
                )
            )

            for a in pm.get("assets", []):
                all_asset_ids.append(a["id"])
                all_asset_reports.append(
                    AssetReport(
                        symbol=a["symbol"],
                        name=a.get("name") or a["symbol"],
                        asset_type=a["asset_type"],
                        quantity=a["quantity"],
                        avg_buy_price=a["avg_buy_price"],
                        current_price=a.get("current_price") or a["avg_buy_price"],
                        total_invested=a["total_invested"],
                        current_value=a["current_value"],
                        gain_loss=a["gain_loss"],
                        gain_loss_percent=a["gain_loss_percent"],
                        breakeven_price=a.get("breakeven_price"),
                        risk_weight=a.get("risk_weight"),
                        total_fees=a.get("total_fees", 0.0),
                        exchange=a.get("exchange"),
                    )
                )

        # 3. Transaction list (direct query — metrics_service doesn't expose rows)
        tax_transactions: List[TaxTransaction] = []
        if all_asset_ids:
            trans_query = select(Transaction).where(Transaction.asset_id.in_(all_asset_ids))
            if year:
                trans_query = trans_query.where(
                    Transaction.executed_at >= datetime(year, 1, 1),
                    Transaction.executed_at <= datetime(year, 12, 31, 23, 59, 59),
                )
            trans_query = trans_query.order_by(Transaction.executed_at.desc())
            trans_result = await db.execute(trans_query)
            transactions = trans_result.scalars().all()

            # Build asset id → symbol map from DB
            asset_id_result = await db.execute(select(Asset.id, Asset.symbol).where(Asset.id.in_(all_asset_ids)))
            id_to_symbol = {str(row[0]): row[1] for row in asset_id_result.all()}

            for trans in transactions:
                symbol = id_to_symbol.get(str(trans.asset_id), "")
                if symbol:
                    tax_transactions.append(
                        TaxTransaction(
                            date=trans.executed_at,
                            symbol=symbol,
                            transaction_type=trans.transaction_type.value,
                            quantity=float(trans.quantity),
                            price=float(trans.price),
                            total=float(trans.quantity) * float(trans.price),
                            fee=float(trans.fee) if trans.fee else 0,
                            gain_loss=None,
                        )
                    )

        # 4. Risk metrics (for Excel "Analyse de Risque" sheet)
        # Use the same allocations as the Dashboard (filtered: current_value > 0)
        # and the same days resolution (days=0 → actual days since first tx)
        # to ensure HHI, VaR, etc. match exactly.
        risk_metrics: Dict[str, Any] = {}
        try:
            allocations = [
                {"symbol": a["symbol"], "value": a["current_value"]} for a in dashboard.get("aggregated_assets", [])
            ]
            # Resolve days like the Dashboard does for days=0 ("all time")
            first_tx_result = await db.execute(
                select(func.min(Transaction.executed_at)).where(Transaction.asset_id.in_(all_asset_ids))
            )
            first_tx_date = first_tx_result.scalar()
            if first_tx_date:
                if hasattr(first_tx_date, "tzinfo") and first_tx_date.tzinfo is not None:
                    first_tx_date = first_tx_date.replace(tzinfo=None)
                risk_days = max((datetime.now(timezone.utc) - first_tx_date).days + 1, 7)
            else:
                risk_days = 30

            risk_metrics = await snapshot_service.get_all_risk_metrics(
                db,
                user_id,
                current_value=dashboard["total_value"],
                allocations=allocations,
                days=risk_days,
            )
        except Exception:
            logger.warning("Failed to compute risk metrics for report", exc_info=True)

        # 5. Platform analysis (GROUP BY exchange)
        platform_map: Dict[str, Dict[str, Any]] = {}
        for ar in all_asset_reports:
            ex = ar.exchange or "Autre"
            if ex not in platform_map:
                platform_map[ex] = {
                    "exchange": ex,
                    "total_value": 0.0,
                    "total_invested": 0.0,
                    "total_fees": 0.0,
                    "count": 0,
                }
            platform_map[ex]["total_value"] += ar.current_value
            platform_map[ex]["total_invested"] += ar.total_invested
            platform_map[ex]["total_fees"] += ar.total_fees
            platform_map[ex]["count"] += 1

        platform_analysis = []
        for p in platform_map.values():
            invested = p["total_invested"]
            p["roi_percent"] = ((p["total_value"] - invested) / invested * 100) if invested > 0 else 0
            # `invested` is fee-inclusive (buy fees are baked into the cost basis), so
            # subtracting total_fees again would double-count them. Net P&L is simply the
            # gain over the fee-inclusive cost.
            p["net_pnl"] = p["total_value"] - invested
            platform_analysis.append(p)
        platform_analysis.sort(key=lambda x: x["total_value"], reverse=True)

        # 6. Attribution: Alpha / Beta (BTC) / Protection (Or) / Fixed Income / Munitions
        attribution = {
            "alpha": 0.0,
            "beta": 0.0,
            "protection": 0.0,
            "fixed_income": 0.0,
            "munitions": 0.0,
        }
        for ar in all_asset_reports:
            sym = ar.symbol.upper()
            if is_liquidity(sym):
                attribution["munitions"] += ar.current_value
            elif getattr(ar, "asset_type", "") == "crowdfunding":
                attribution["fixed_income"] += ar.current_value
            elif sym == "BTC":
                attribution["beta"] += ar.current_value
            elif is_safe_haven(sym):
                attribution["protection"] += ar.current_value
            else:
                attribution["alpha"] += ar.current_value

        # 7. Cash flows (deposits are NOT gains)
        # M2: TRANSFER_IN/OUT are NOT capital flows — they just move assets
        # between exchanges. Only BUY = deposit, SELL = withdrawal.
        cash_flows = {"total_deposits": 0.0, "total_withdrawals": 0.0}
        for tx in tax_transactions:
            tt = tx.transaction_type
            if tt == "buy":
                cash_flows["total_deposits"] += tx.total
            elif tt == "sell":
                cash_flows["total_withdrawals"] += tx.total
        cash_flows["net_flows"] = cash_flows["total_deposits"] - cash_flows["total_withdrawals"]

        # 8. Build result
        pnl_data = dashboard.get("pnl_data", {})

        return {
            "portfolios": portfolio_summaries,
            "assets": all_asset_reports,
            "transactions": tax_transactions,
            "summary": {
                "total_value": dashboard["total_value"],
                "total_invested": dashboard["total_invested"],
                "gain_loss": dashboard.get("net_gain_loss", 0),
                "gain_loss_percent": dashboard.get("net_gain_loss_percent", 0),
            },
            "pnl_data": pnl_data,
            "risk_metrics": risk_metrics,
            "platform_analysis": platform_analysis,
            "attribution": attribution,
            "cash_flows": cash_flows,
            "generated_at": datetime.now(timezone.utc),
            "year": year,
        }

    # ── Tax 2086 computation (FIFO-aligned) ─────────────────────────

    async def generate_tax_report_2086(
        self,
        db: AsyncSession,
        user_id: str,
        year: int,
    ) -> bytes:
        """Generate French tax report 2086 for crypto assets with real capital gains."""
        tax = await self.compute_tax_2086(db, user_id, year)

        buffer = io.BytesIO()
        doc = _build_doc_with_footer(
            buffer,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("T", parent=styles["Heading1"], fontSize=20, spaceAfter=20, textColor=_BLUE)
        heading_style = ParagraphStyle(
            "H",
            parent=styles["Heading2"],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=8,
            textColor=_DARK_BLUE,
        )
        info_style = ParagraphStyle(
            "I",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.HexColor("#64748b"),
        )
        small_style = ParagraphStyle(
            "S",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor("#64748b"),
        )

        elements = []

        # ── Title ───────────────────────────────────────────────────
        elements.append(Paragraph(f"Déclaration Fiscale Crypto — Année {year}", title_style))
        elements.append(Paragraph("Formulaire 2086 — Plus-values sur actifs numériques", heading_style))
        elements.append(
            Paragraph(
                f"Document généré le {_now_paris().strftime('%d/%m/%Y')} — À usage indicatif uniquement",
                info_style,
            )
        )
        elements.append(Spacer(1, 20))

        # ── 1. Résumé des cessions ──────────────────────────────────
        elements.append(Paragraph("1. Résumé des cessions", heading_style))

        summary_rows = [
            ["Description", "Montant"],
            ["Nombre de cessions", str(tax.nb_cessions)],
            ["  dont court terme (< 2 ans)", str(tax.nb_court_terme)],
            ["  dont long terme (≥ 2 ans)", str(tax.nb_long_terme)],
            ["Prix total de cession", _money(tax.total_cessions)],
            ["Fraction d'acquisition (PMP)", _money(tax.total_acquisitions_fraction)],
            ["Total plus-values", _money(tax.total_plus_values)],
            ["Total moins-values", _money(tax.total_moins_values)],
            ["Plus-value nette imposable", _money(tax.net_plus_value)],
        ]

        t = Table(summary_rows, colWidths=[10 * cm, 5 * cm])
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            # Highlight net PV row
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        if tax.net_plus_value >= 0:
            style_cmds.append(("TEXTCOLOR", (1, -1), (1, -1), _GREEN))
        else:
            style_cmds.append(("TEXTCOLOR", (1, -1), (1, -1), _RED))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # ── 2. Estimation fiscale ───────────────────────────────────
        elements.append(Paragraph("2. Estimation de l'imposition (PFU — Flat Tax 30%)", heading_style))

        tax_rows = [
            ["Composante", "Taux", "Montant"],
            ["Plus-value nette imposable", "", _money(max(0, tax.net_plus_value))],
            ["Impôt sur le revenu", "12,8%", _money(tax.ir_12_8)],
            ["Prélèvements sociaux", "17,2%", _money(tax.ps_17_2)],
            ["Total PFU (Flat Tax)", "30,0%", _money(tax.flat_tax_30)],
        ]

        t = Table(tax_rows, colWidths=[8 * cm, 3 * cm, 4 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
                ]
            )
        )
        elements.append(t)

        if tax.net_plus_value <= 0:
            elements.append(Spacer(1, 5))
            elements.append(
                Paragraph(
                    "Aucune imposition : la plus-value nette est négative ou nulle.",
                    info_style,
                )
            )

        elements.append(Spacer(1, 20))

        # ── 3. Détail des cessions ──────────────────────────────────
        if tax.events:
            elements.append(Paragraph("3. Détail des cessions", heading_style))
            elements.append(
                Paragraph(
                    "Calcul selon la formule 2086 : PV = Prix cession − (Coût acquisition global × Prix cession / Valeur portefeuille)",
                    small_style,
                )
            )
            elements.append(Spacer(1, 8))

            header = [
                "Date",
                "Actif",
                "Type",
                "Qté",
                "Prix cession",
                "Fraction acq.",
                "PV/MV",
                "Durée",
            ]
            rows = [header]
            for e in sorted(tax.events, key=lambda x: x.date):
                type_label = "Conv." if e.event_type == "conversion_out" else "Vente"
                duration = "≥2 ans" if e.holding_period == "long_terme" else "<2 ans"
                rows.append(
                    [
                        e.date.strftime("%d/%m/%Y"),
                        e.symbol,
                        type_label,
                        f"{e.quantity:,.6f}",
                        _money(e.cession_price),
                        _money(e.acquisition_fraction),
                        _money(e.gain_loss),
                        duration,
                    ]
                )

            t = Table(
                rows,
                colWidths=[
                    2 * cm,
                    1.6 * cm,
                    1.4 * cm,
                    2 * cm,
                    2.4 * cm,
                    2.4 * cm,
                    2.2 * cm,
                    1.6 * cm,
                ],
            )
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_BG]),
            ]
            # Color gain/loss column
            for i, e in enumerate(sorted(tax.events, key=lambda x: x.date), 1):
                c = _GREEN if e.gain_loss >= 0 else _RED
                style_cmds.append(("TEXTCOLOR", (6, i), (6, i), c))
            t.setStyle(TableStyle(style_cmds))
            elements.append(t)
        else:
            elements.append(
                Paragraph(
                    "Aucune cession d'actifs numériques enregistrée pour cette année.",
                    info_style,
                )
            )

        elements.append(Spacer(1, 30))

        # ── 4. Mentions légales ─────────────────────────────────────
        elements.append(Paragraph("Mentions importantes", heading_style))
        elements.append(
            Paragraph(
                "Ce document est fourni à titre indicatif uniquement et ne constitue pas un conseil fiscal. "
                "Veuillez consulter un professionnel pour votre déclaration officielle. "
                "Les plus-values sur actifs numériques sont soumises au prélèvement forfaitaire unique (PFU) de 30% "
                "(12,8% d'impôt sur le revenu + 17,2% de prélèvements sociaux) ou au barème progressif de l'impôt sur le revenu sur option. "
                "Le calcul utilise la méthode du Prix Moyen Pondéré (PMP) global du portefeuille conformément à l'article 150 VH bis du CGI. "
                "La valeur du portefeuille au moment de chaque cession est estimée à partir des données disponibles.",
                info_style,
            )
        )

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Tax Excel ───────────────────────────────────────────────────

    async def generate_tax_excel(
        self,
        db: AsyncSession,
        user_id: str,
        year: int,
    ) -> bytes:
        """Generate Excel file for tax reporting with full 2086 calculations."""
        tax = await self.compute_tax_2086(db, user_id, year)

        wb = Workbook()

        # ── Résumé sheet ────────────────────────────────────────────
        ws = wb.active
        ws.title = f"Résumé {year}"

        ws["A1"] = f"Déclaration Fiscale Crypto — Année {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Généré le {_now_paris().strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Summary data
        ws["A4"] = "Résumé des cessions"
        ws["A4"].font = Font(bold=True, size=12)

        summary_items = [
            ("Nombre de cessions", tax.nb_cessions, None),
            ("  dont court terme (< 2 ans)", tax.nb_court_terme, None),
            ("  dont long terme (≥ 2 ans)", tax.nb_long_terme, None),
            ("Prix total de cession", tax.total_cessions, _XL_MONEY),
            (
                "Fraction d'acquisition (PMP)",
                tax.total_acquisitions_fraction,
                _XL_MONEY,
            ),
            ("Total plus-values", tax.total_plus_values, _XL_MONEY),
            ("Total moins-values", tax.total_moins_values, _XL_MONEY),
            ("Plus-value nette imposable", tax.net_plus_value, _XL_MONEY),
        ]
        for i, (label, value, fmt) in enumerate(summary_items, 5):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            if fmt:
                ws[f"B{i}"].number_format = fmt

        # Tax estimation
        row = 5 + len(summary_items) + 1
        ws[f"A{row}"] = "Estimation fiscale (PFU 30%)"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        tax_items = [
            ("Impôt sur le revenu (12,8%)", tax.ir_12_8),
            ("Prélèvements sociaux (17,2%)", tax.ps_17_2),
            ("Total Flat Tax (30%)", tax.flat_tax_30),
        ]
        for i, (label, value) in enumerate(tax_items, row + 1):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].number_format = _XL_MONEY
            if label.startswith("Total"):
                ws[f"A{i}"].font = Font(bold=True)
                ws[f"B{i}"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 18

        # ── Cessions sheet ──────────────────────────────────────────
        ws2 = wb.create_sheet("Cessions")
        headers = [
            "Date",
            "Actif",
            "Type",
            "Quantité",
            "Prix unitaire",
            "Prix cession",
            "Valeur portefeuille",
            "Coût acquisition",
            "Fraction acq.",
            "PV/MV",
            "Durée",
        ]
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = _XL_HEADER_FONT
            c.fill = _XL_HEADER_FILL
            c.border = _XL_BORDER

        for row, e in enumerate(sorted(tax.events, key=lambda x: x.date), 2):
            type_label = "Conversion" if e.event_type == "conversion_out" else "Vente"
            duration = "≥ 2 ans" if e.holding_period == "long_terme" else "< 2 ans"
            vals = [
                e.date.strftime("%d/%m/%Y"),
                e.symbol,
                type_label,
                e.quantity,
                e.unit_price,
                e.cession_price,
                e.portfolio_value,
                e.total_acquisition_cost,
                e.acquisition_fraction,
                e.gain_loss,
                duration,
            ]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=row, column=col, value=v)
                c.border = _XL_BORDER
                if col in (5, 6, 7, 8, 9, 10):
                    c.number_format = _XL_MONEY
                if col == 10:
                    c.font = Font(color="16a34a" if e.gain_loss >= 0 else "dc2626")

        for col in range(1, 12):
            ws2.column_dimensions[get_column_letter(col)].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Transactions PDF (dedicated) ────────────────────────────────


# Singleton
report_service = ReportService()
