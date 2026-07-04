"""Transactions report rendering (PDF / Excel / CSV), extracted from report_service."""

from __future__ import annotations

import io
import logging
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

from app.services.report_common import (
    _BLUE,
    _BORDER_COLOR,
    _DARK_BLUE,
    _LIGHT_BG,
    _XL_BORDER,
    _XL_HEADER_FILL,
    _XL_HEADER_FONT,
    _XL_MONEY,
    _build_doc_with_footer,
    _fmt_qty,
    _format_paris_dt,
    _money,
)

logger = logging.getLogger(__name__)


class TransactionsReportMixin:
    """Mixed into ReportService."""

    def generate_transactions_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate a dedicated PDF with all transactions."""
        transactions = data.get("transactions", [])

        buffer = io.BytesIO()
        doc = _build_doc_with_footer(
            buffer,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("T", parent=styles["Heading1"], fontSize=20, spaceAfter=20, textColor=_BLUE)
        heading_style = ParagraphStyle(
            "H",
            parent=styles["Heading2"],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=8,
            textColor=_DARK_BLUE,
        )
        normal_style = styles["Normal"]
        elements = []

        elements.append(Paragraph("Historique des Transactions", title_style))
        year = data.get("year")
        period = f"Année {year}" if year else "Toutes les années"
        elements.append(Paragraph(f"{period} — Généré le {_format_paris_dt()}", normal_style))
        elements.append(Spacer(1, 15))

        # Summary
        elements.append(Paragraph("Résumé", heading_style))
        nb = len(transactions)
        total_volume = sum(t.total for t in transactions)
        total_fees = sum(t.fee for t in transactions)
        summary_rows = [
            ["Nombre de transactions", str(nb)],
            ["Volume total", _money(total_volume)],
            ["Total frais", _money(total_fees)],
        ]
        t = Table(summary_rows, colWidths=[8 * cm, 5 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("BACKGROUND", (0, 0), (-1, -1), _LIGHT_BG),
                ]
            )
        )
        elements.append(t)
        elements.append(Spacer(1, 15))

        # Transaction table
        if transactions:
            elements.append(Paragraph("Détail", heading_style))
            _TYPE_MAP = {
                "buy": "Achat",
                "sell": "Vente",
                "transfer_in": "Transfert ↓",
                "transfer_out": "Transfert ↑",
                "conversion_in": "Conv. ↓",
                "conversion_out": "Conv. ↑",
                "airdrop": "Airdrop",
                "staking_reward": "Reward",
                "fee": "Frais",
                "dividend": "Dividende",
                "interest": "Intérêt",
                "staking": "Staking",
                "unstaking": "Unstaking",
            }
            rows = [
                [
                    "Date",
                    "Type",
                    "Actif",
                    "Quantité",
                    "Prix Unitaire",
                    "Valeur Totale",
                    "Frais",
                ]
            ]
            for tx in transactions:
                rows.append(
                    [
                        tx.date.strftime("%d/%m/%Y") if tx.date else "",
                        _TYPE_MAP.get(tx.transaction_type, tx.transaction_type),
                        tx.symbol,
                        _fmt_qty(tx.quantity),
                        _money(tx.price),
                        _money(tx.total),
                        _money(tx.fee),
                    ]
                )
            t = Table(
                rows,
                colWidths=[
                    2.2 * cm,
                    2.2 * cm,
                    2 * cm,
                    2.4 * cm,
                    2.8 * cm,
                    2.8 * cm,
                    2 * cm,
                ],
            )
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 1, _BORDER_COLOR),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_BG]),
                    ]
                )
            )
            elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Transactions Excel ───────────────────────────────────────────

    def generate_transactions_excel(self, data: Dict[str, Any]) -> bytes:
        """Generate an Excel file with full transaction history."""
        transactions = data.get("transactions", [])
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        _TYPE_MAP = {
            "buy": "Achat",
            "sell": "Vente",
            "transfer_in": "Transfert entrant",
            "transfer_out": "Transfert sortant",
            "conversion_in": "Conversion entrante",
            "conversion_out": "Conversion sortante",
            "airdrop": "Airdrop",
            "staking_reward": "Reward",
            "fee": "Frais",
            "dividend": "Dividende",
            "interest": "Intérêt",
            "staking": "Staking",
            "unstaking": "Unstaking",
        }

        headers = [
            "Date",
            "Type",
            "Actif",
            "Quantité",
            "Prix Unitaire (EUR)",
            "Valeur Totale (EUR)",
            "Frais (EUR)",
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = _XL_HEADER_FONT
            c.fill = _XL_HEADER_FILL
            c.border = _XL_BORDER

        for row, tx in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=tx.date.strftime("%d/%m/%Y") if tx.date else "").border = _XL_BORDER
            ws.cell(
                row=row,
                column=2,
                value=_TYPE_MAP.get(tx.transaction_type, tx.transaction_type),
            ).border = _XL_BORDER
            ws.cell(row=row, column=3, value=tx.symbol).border = _XL_BORDER
            c = ws.cell(row=row, column=4, value=tx.quantity)
            c.number_format = "0.000000"
            c.border = _XL_BORDER
            ws.cell(row=row, column=5, value=tx.price).number_format = _XL_MONEY
            ws.cell(row=row, column=5).border = _XL_BORDER
            ws.cell(row=row, column=6, value=tx.total).number_format = _XL_MONEY
            ws.cell(row=row, column=6).border = _XL_BORDER
            ws.cell(row=row, column=7, value=tx.fee).number_format = _XL_MONEY
            ws.cell(row=row, column=7).border = _XL_BORDER

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ── Transactions CSV ─────────────────────────────────────────────

    def generate_transactions_csv(self, data: Dict[str, Any]) -> bytes:
        """Generate a CSV file with full transaction history."""
        import csv

        transactions = data.get("transactions", [])

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow(
            [
                "Date",
                "Type",
                "Actif",
                "Quantité",
                "Prix Unitaire (EUR)",
                "Valeur Totale (EUR)",
                "Frais (EUR)",
            ]
        )

        for tx in transactions:
            writer.writerow(
                [
                    tx.date.strftime("%d/%m/%Y") if tx.date else "",
                    tx.transaction_type,
                    tx.symbol,
                    f"{tx.quantity:.6f}",
                    f"{tx.price:.2f}",
                    f"{tx.total:.2f}",
                    f"{tx.fee:.2f}",
                ]
            )

        return buffer.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    # ── Rebalancing Report ────────────────────────────────────────────
